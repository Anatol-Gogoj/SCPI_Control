#!/usr/bin/env python3
"""Headless tests for the battery-cycler processing (issue #31).

Builds a tiny synthetic tester workbook (Info + 2 Detail sheets with
Mandarin headers and both proper and GBK-garbled status labels + Cycle
summary) and checks the translation/concatenation pipeline. Parity with
the original standalone tool was verified separately against the real
sample export (see PR #68). Skips cleanly if pandas/openpyxl are absent.

Run: .venv/bin/python tests/test_battery_process.py
"""
# Runnable from anywhere: put the repo root (one level up) on sys.path
# so the app modules import when this file is executed directly.
import os as _os
import sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(
    _os.path.abspath(__file__))))

import shutil
import tempfile

from battery_process import (HEADER_MAP, STATUS_MAP, axis_label,
                             deps_available, parse_cycle_selection)

_DEPS_OK, _DEPS_WHY = deps_available()

_ZH_HEADERS = ["状态", "循环", "电流(mA)", "电压(mV)", "容量(mAH)",
               "相对时间(秒)"]
_GARBLED_CC = "恒流充电".encode("gbk").decode("latin-1")


def _make_workbook(path):
    from openpyxl import Workbook
    wb = Workbook()
    info = wb.active
    info.title = "Info"
    info.append(["device", "test-42"])
    d1 = wb.create_sheet("Detail_1")
    d1.append(_ZH_HEADERS)
    d1.append(["恒流充电", 1, 100.0, 3300.0, 0.5, 10.0])
    d1.append(["搁置", 1, 0.0, 3290.0, 0.5, 20.0])
    d2 = wb.create_sheet("Detail_2")
    d2.append(_ZH_HEADERS)
    d2.append([_GARBLED_CC, 2, 100.0, 3400.0, 1.0, 30.0])
    d2.append(["恒流放电", 2, -100.0, 3200.0, 0.8, 40.0])
    cyc = wb.create_sheet("Cycle")
    cyc.append(["cycle summary -- must be skipped"])
    wb.save(path)


def test_parse_cycle_selection():
    assert parse_cycle_selection("1,2,3") == [1, 2, 3]
    assert parse_cycle_selection("1-4") == [1, 2, 3, 4]
    assert parse_cycle_selection("7, 2-4 ,junk") == [2, 3, 4, 7]
    assert parse_cycle_selection("") is None
    assert parse_cycle_selection(None) is None
    assert parse_cycle_selection("e.g. 1,2,3 or 1-5") is None
    assert parse_cycle_selection("nonsense") is None


def test_axis_label():
    assert axis_label("relative_time_s", "min") == "relative_time (min)"
    assert axis_label("cycle_time_s", "h") == "cycle_time (h)"
    assert axis_label("voltage_V", "min") == "voltage_V"


def test_status_map_covers_garbled_forms():
    assert STATUS_MAP["恒流充电"] == "CC_charge"
    assert STATUS_MAP[_GARBLED_CC] == "CC_charge", \
        "GBK-as-Latin-1 garbled labels must translate too"
    assert STATUS_MAP["静置"] == "rest" and STATUS_MAP["搁置"] == "rest"


def test_load_and_process_synthetic():
    if not _DEPS_OK:
        print(f"SKIP (deps missing: {_DEPS_WHY})")
        return
    from battery_process import load_and_process
    d = tempfile.mkdtemp(prefix="battery_")
    try:
        path = os.path.join(d, "export.xlsx")
        _make_workbook(path)
        df = load_and_process(path)
        # both Detail sheets concatenated, Info/Cycle skipped
        assert len(df) == 4, len(df)
        # headers translated
        for col in ("status", "cycle", "current_mA", "voltage_mV",
                    "capacity_mAh", "relative_time_s"):
            assert col in df.columns, df.columns.tolist()
        # statuses translated, incl. the garbled one
        assert df["status"].tolist() == ["CC_charge", "rest",
                                         "CC_charge", "CC_discharge"]
        # derived volts
        assert abs(df["voltage_V"].iloc[0] - 3.3) < 1e-9
        # numerics are numeric
        assert df["cycle"].tolist() == [1, 1, 2, 2]
    finally:
        shutil.rmtree(d)


def test_load_rejects_wrong_shape():
    if not _DEPS_OK:
        print(f"SKIP (deps missing: {_DEPS_WHY})")
        return
    from openpyxl import Workbook
    from battery_process import load_and_process
    d = tempfile.mkdtemp(prefix="battery_")
    try:
        path = os.path.join(d, "two_sheets.xlsx")
        wb = Workbook()
        wb.active.title = "Info"
        wb.create_sheet("Cycle")
        wb.save(path)
        try:
            load_and_process(path)
            assert False, "must raise on <3 sheets"
        except ValueError:
            pass
    finally:
        shutil.rmtree(d)


import os  # noqa: E402  (used by tests above)


def _run():
    fns = [v for k, v in sorted(globals().items()) if k.startswith('test_')]
    for fn in fns:
        fn()
        print(f"ok  {fn.__name__}")
    print(f"\n{len(fns)} tests passed")


if __name__ == '__main__':
    _run()
